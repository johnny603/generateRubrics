"""
make_demo_files.py - Generates dummy Excel rubrics for the samples/ folder.
Run this once before using generate_rubrics.py to test the tool.
"""

import os
import sys

# Try to use openpyxl; if not installed, fallback to CSV.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    import csv

def create_excel(filename, assignment_num):
    wb = Workbook()
    ws = wb.active
    ws.title = f"ASMT {assignment_num:02d}"

    # Header
    ws['A1'] = f"Assignment {assignment_num:02d} Rubric (DEMO)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A3'] = "Criteria"
    ws['B3'] = "Max Points"
    ws['C3'] = "Points Received"
    ws['D3'] = "Feedback"
    ws['E3'] = "Notes"

    # Fake rows
    criteria = [
        ("Part A - Analysis", 5),
        ("Part B - Code Quality", 10),
        ("Part C - Output", 5),
        ("Part D - Documentation", 5),
        ("Part E - Extra Credit", 3),
    ]
    row = 4
    for crit, pts in criteria:
        ws.cell(row, 1, crit)
        ws.cell(row, 2, pts)
        ws.cell(row, 3, "")  # Points Received
        ws.cell(row, 4, "")  # Feedback
        ws.cell(row, 5, "")  # Notes
        row += 1

    # Total
    ws.cell(row, 1, "TOTAL")
    ws.cell(row, 2, f"=SUM(B4:B{row-1})")
    ws.cell(row, 1).font = Font(bold=True)

    # Adjust column widths
    for col in 'ABCDE':
        ws.column_dimensions[col].width = 20

    wb.save(filename)
    print(f"Created Excel demo rubric: {filename}")

def create_csv(filename, assignment_num):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([f"Assignment {assignment_num:02d} Rubric (DEMO)"])
        writer.writerow([])
        writer.writerow(["Criteria", "Max Points", "Points Received", "Feedback", "Notes"])
        writer.writerow(["Part A - Analysis", 5, "", "", ""])
        writer.writerow(["Part B - Code Quality", 10, "", "", ""])
        writer.writerow(["Part C - Output", 5, "", "", ""])
        writer.writerow(["Part D - Documentation", 5, "", "", ""])
        writer.writerow(["Part E - Extra Credit", 3, "", "", ""])
        writer.writerow(["TOTAL", "=SUM(B4:B8)", "", "", ""])
    print(f"Created CSV fallback (open in Excel): {filename}")

def main():
    # Ensure we are in the 'samples' folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    for num in [1, 2]:
        filename = f"Assignment-{num:02d}-Rubric.xlsx" if HAS_OPENPYXL else f"Assignment-{num:02d}-Rubric.csv"
        if os.path.exists(filename):
            print(f"Skipping {filename} (already exists)")
            continue
        if HAS_OPENPYXL:
            create_excel(filename, num)
        else:
            create_csv(filename, num)

    print("\nDemo files ready. Now you can run:")
    print("  python generate_rubrics.py -inputFile samples/studentNames.txt -fileToCopy samples/Assignment-01-Rubric.xlsx --assignment 01")
    print("or use the wrapper scripts.")

if __name__ == "__main__":
    main()
EOF
